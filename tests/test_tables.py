"""Tests for the camelot-py-backed PDF to Excel operation (the `extras` group).

Skips if camelot-py isn't installed, same adaptive pattern as test_convert.py --
this machine has it installed, so the real detection path is exercised.
"""

from __future__ import annotations

import importlib.util
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pypdf import PdfWriter
from pypdf.generic import DecodedStreamObject, DictionaryObject, NameObject

from quire.core.registry import REGISTRY, load_operations

load_operations()

HAS_CAMELOT = importlib.util.find_spec("camelot") is not None


def _font_resources(writer: PdfWriter) -> DictionaryObject:
    font = DictionaryObject(
        {
            NameObject("/Type"): NameObject("/Font"),
            NameObject("/Subtype"): NameObject("/Type1"),
            NameObject("/BaseFont"): NameObject("/Helvetica"),
        }
    )
    return DictionaryObject(
        {NameObject("/Font"): DictionaryObject({NameObject("/Helv"): writer._add_object(font)})}
    )


@pytest.fixture
def ruled_table_pdf(tmp_path: Path) -> Path:
    """A single page with one hand-drawn 2x2 ruled table -- lines + text,
    built by hand since camelot needs real ruling to detect a table."""
    writer = PdfWriter()
    page = writer.add_blank_page(width=300, height=200)
    ops = b"""
    q 1 w
    0 0 0 RG
    50 100 200 60 re S
    150 100 m 150 160 l S
    50 130 m 250 130 l S
    BT /Helv 10 Tf 60 145 Td (Name) Tj ET
    BT /Helv 10 Tf 160 145 Td (Score) Tj ET
    BT /Helv 10 Tf 60 112 Td (Alice) Tj ET
    BT /Helv 10 Tf 160 112 Td (92) Tj ET
    Q
    """
    content = DecodedStreamObject()
    content.set_data(ops)
    page[NameObject("/Contents")] = writer._add_object(content)
    page[NameObject("/Resources")] = _font_resources(writer)
    target = tmp_path / "table.pdf"
    with target.open("wb") as fh:
        writer.write(fh)
    return target


@pytest.fixture
def sample(tmp_path: Path) -> Path:
    writer = PdfWriter()
    writer.add_blank_page(width=200, height=200)
    target = tmp_path / "sample.pdf"
    with target.open("wb") as fh:
        writer.write(fh)
    return target


def test_pdf_to_excel_reports_missing_extra(sample, tmp_path):
    if HAS_CAMELOT:
        pytest.skip("camelot-py is installed; see test_pdf_to_excel_detects_table instead")
    with pytest.raises(ValueError, match="extras"):
        REGISTRY["pdf_to_excel"].run([sample], tmp_path / "o")


def test_pdf_to_excel_detects_table(ruled_table_pdf, tmp_path):
    if not HAS_CAMELOT:
        pytest.skip("camelot-py not installed")
    out = REGISTRY["pdf_to_excel"].run([ruled_table_pdf], tmp_path / "o")[0]
    assert out.suffix == ".xlsx"
    workbook = load_workbook(out)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0] == ("Name", "Score")
    assert rows[1] == ("Alice", "92")


def test_pdf_to_excel_rejects_document_without_tables(sample, tmp_path):
    if not HAS_CAMELOT:
        pytest.skip("camelot-py not installed")
    with pytest.raises(ValueError):
        REGISTRY["pdf_to_excel"].run([sample], tmp_path / "o")
